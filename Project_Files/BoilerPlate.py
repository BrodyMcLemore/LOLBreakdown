from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from datetime import datetime


def MakeBoilerPlateCells(worksheet, startRow, matchDetail,CellValues):
    ws = worksheet
    BoilerPlateCellValues = CellValues
    ws.merge_cells(start_row=startRow, start_column=1, end_row=startRow, end_column=len(BoilerPlateCellValues))
    RedTeamTitleCell = ws.cell(startRow, 1)
    RedTeamTitleCell.value = "Red Team"
    RedTeamTitleCell.alignment = Alignment(horizontal='center')
    RedTeamTitleCell.font = Font(size=14,bold=True)

    ws.merge_cells(start_row=startRow, start_column=len(BoilerPlateCellValues)+2,
                   end_row=startRow, end_column=(len(BoilerPlateCellValues)*2)+1)
    BlueTeamTitleCell = ws.cell(startRow, len(BoilerPlateCellValues)+2)
    BlueTeamTitleCell.value = "Blue Team"
    BlueTeamTitleCell.alignment = Alignment(horizontal='center')
    BlueTeamTitleCell.font = Font(size=14,bold=True)

    ws.merge_cells(start_row=startRow+1, start_column=1,
                   end_row=startRow+1, end_column=(len(BoilerPlateCellValues)*2)+1)
    GameDate = ws.cell(startRow+1, 1)
    GameDate.value = datetime.fromtimestamp(
        (matchDetail["info"]['gameStartTimestamp'] / float(1000))).strftime("%m/%d/%Y, %H:%M:%S")
    GameDate.alignment = Alignment(horizontal='center')

    # Red Team
    for i in range(len(BoilerPlateCellValues)):
        cell = ws.cell(startRow+2, i+1)
        cell.value = BoilerPlateCellValues[i]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    # Blue Team
    for i in range(len(BoilerPlateCellValues)):
        cell = ws.cell(startRow+2, i+len(BoilerPlateCellValues)+2)
        cell.value = BoilerPlateCellValues[i]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    # Divider
    for i in range(7):
        cell = ws.cell(startRow+i+1, len(BoilerPlateCellValues)+1)
        cell.fill = PatternFill("solid", start_color="000000")
