from Project_Files.BoilerPlate import MakeBoilerPlateCells
from openpyxl import Workbook
from riotwatcher import LolWatcher
import os
from Project_Files.DictCalls import *
from tkinter import ttk
from tkinter.filedialog import asksaveasfilename


def SaveBook(Workbook, PlayerName):
    wb = Workbook
    savedFile = asksaveasfilename(
        filetypes=[("Excel File", ".xlsx")],
        defaultextension=".xlsx",
        title="Save LOLBreakdown Output",
        initialfile=PlayerName,
        confirmoverwrite=False
    )

    if savedFile:
        try:
            wb.save(savedFile)
            return "Success"
        except:
            return "Fail"
    else:
        return "Fail"

def OutputGames(Region, MatchList, GameCount, Row, CallValues, Watcher, Workbook, PlayerName):
    watcher = Watcher
    wb = Workbook
    wb.create_sheet(str(PlayerName))
    while GameCount <= 17:
        MatchDetail = watcher.match.by_id(Region, MatchList[GameCount])
        if MatchDetail["info"]["gameMode"] == "CLASSIC":
            MakeBoilerPlateCells(wb[str(PlayerName)],
                                 Row, MatchDetail, CallValues)
            OutputRed(wb[str(PlayerName)], MatchDetail,
                      Row, CallValues)
            OutputBlue(wb[str(PlayerName)], MatchDetail,
                       Row, CallValues)
            Row += 8
        GameCount += 1
    del wb["Sheet"]


def OutputRed(worksheet, MatchDetail, startRow, CallValues):
    ws = worksheet
    StartWorkRow = startRow + 3
    for row in range(5):
        for col in range(len(CallValues)):
            if CallValues[col] == "KDA":
                CalledValue = ws.cell(StartWorkRow+row, col+1)
                Kills = str(MatchDetail["info"]['participants'][row]["kills"])
                Deaths = str(MatchDetail["info"]
                             ['participants'][row]["deaths"])
                Assists = str(MatchDetail["info"]
                              ['participants'][row]["assists"])
                CalledValue.value = Kills + "/" + Deaths + "/" + Assists

            elif CallValues[col] == "inventory":
                Items = ["item0", "item1", "item2", "item3", "item4", "item5"]
                ItemNames = []
                ItemsDict = GetItemDict()
                CalledValue = ws.cell(StartWorkRow+row, col+1)
                for i in range(5):
                    if str(MatchDetail["info"]['participants'][row][Items[i]]) != "0":
                        ItemNames.append(
                            ItemsDict[str(MatchDetail["info"]['participants'][row][Items[i]])])
                CalledValue.value = str(str(", ".join(ItemNames)))

            elif CallValues[col] == "summonerSpells":
                SummonerSpells = GetSummonerSpellDict()
                CalledValue = ws.cell(StartWorkRow+row, col+1)
                CalledValue.value = str(SummonerSpells[str(MatchDetail["info"]['participants'][row]["summoner1Id"])]) + " and " + str(
                    SummonerSpells[str(MatchDetail["info"]['participants'][row+5]["summoner2Id"])])

            elif CallValues[col] == "runes":
                RunesDict = GetRuneDict()
                CalledValue = ws.cell(StartWorkRow+row, col+1)
                Runes = []
                for y in range(3):
                    for x in range(4):
                        try:
                            Runes.append(RunesDict[str(
                                MatchDetail["info"]["participants"][row]["perks"]["styles"][y]["selections"][x]["perk"])])
                        except:
                            continue
                CalledValue.value = str(Runes[0] + " and " + Runes[4])

            elif CallValues[col] == "creepScore":
                CalledValue = ws.cell(StartWorkRow+row, col+1)
                MinionsKilled = MatchDetail["info"]['participants'][row]["totalMinionsKilled"]
                NeutralMinionsKilled = MatchDetail["info"]['participants'][row]["neutralMinionsKilled"]
                CalledValue.value = str(int(MinionsKilled+NeutralMinionsKilled))
            
            elif CallValues[col] == "gameResult":
                CalledValue = ws.cell(StartWorkRow+row, col+1)
                if MatchDetail["info"]['participants'][row]["win"]:
                   CalledValue.value = "Won"
                else:
                    CalledValue.value = "Lost" 

            else:
                CalledValue = ws.cell(StartWorkRow+row, col+1)
                CalledValue.value = MatchDetail["info"]['participants'][row][CallValues[col]]


def OutputBlue(worksheet, MatchDetail, startRow, CallValues):
    ws = worksheet
    StartWorkRow = startRow + 3
    for row in range(5):
        for col in range(len(CallValues)):
            if CallValues[col] == "KDA":
                CalledValue = ws.cell(StartWorkRow+row, col+2+len(CallValues))
                Kills = str(MatchDetail["info"]
                            ['participants'][row+5]["kills"])
                Deaths = str(MatchDetail["info"]
                             ['participants'][row+5]["deaths"])
                Assists = str(MatchDetail["info"]
                              ['participants'][row+5]["assists"])
                CalledValue.value = Kills + "/" + Deaths + "/" + Assists

            elif CallValues[col] == "inventory":
                Items = ["item0", "item1", "item2", "item3", "item4", "item5"]
                ItemNames = []
                ItemsDict = GetItemDict()
                CalledValue = ws.cell(StartWorkRow+row, col+2+len(CallValues))
                for i in range(5):
                    if str(MatchDetail["info"]['participants'][row+5][Items[i]]) != "0":
                        ItemNames.append(
                            ItemsDict[str(MatchDetail["info"]['participants'][row+5][Items[i]])])
                CalledValue.value = str(str(", ".join(ItemNames)))

            elif CallValues[col] == "summonerSpells":
                SummonerSpells = GetSummonerSpellDict()
                CalledValue = ws.cell(StartWorkRow+row, col+2+len(CallValues))
                CalledValue.value = str(SummonerSpells[str(MatchDetail["info"]['participants'][row+5]["summoner1Id"])]) + " and " + str(
                    SummonerSpells[str(MatchDetail["info"]['participants'][row+5]["summoner2Id"])])

            elif CallValues[col] == "runes":
                RunesDict = GetRuneDict()
                CalledValue = ws.cell(StartWorkRow+row, col+2+len(CallValues))
                Runes = []
                for y in range(3):
                    for x in range(4):
                        try:
                            Runes.append(RunesDict[str(
                                MatchDetail["info"]["participants"][row+5]["perks"]["styles"][y]["selections"][x]["perk"])])
                        except:
                            continue
                CalledValue.value = str(Runes[0] + " and " + Runes[4])
            
            elif CallValues[col] == "creepScore":
                CalledValue = ws.cell(StartWorkRow+row, col+2+len(CallValues))
                MinionsKilled = MatchDetail["info"]['participants'][row+5]["totalMinionsKilled"]
                NeutralMinionsKilled = MatchDetail["info"]['participants'][row+5]["neutralMinionsKilled"]
                CalledValue.value = str(int(MinionsKilled+NeutralMinionsKilled))
            
            elif CallValues[col] == "gameResult":
                CalledValue = ws.cell(StartWorkRow+row, col+2+len(CallValues))
                if MatchDetail["info"]['participants'][row+5]["win"]:
                   CalledValue.value = "Won"
                else:
                    CalledValue.value = "Lost"
            
            else:
                CalledValue = ws.cell(StartWorkRow+row, col+2+len(CallValues))
                CalledValue.value = MatchDetail["info"]['participants'][row + 5][CallValues[col]]
