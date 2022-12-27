from riotwatcher import LolWatcher
from openpyxl import Workbook
from datetime import datetime
import os
from ProjectFiles.BoilerPlate import MakeBoilerPlateCells
from time import sleep


def OutputRed(worksheet,matchDetail,startRow,CallValues):
    ws = worksheet
    StartWorkRow = startRow + 3

    for row in range(5):
        for col in range(len(CallValues)):
            if CallValues[col] != "kills":
                #ws.cell(Row,Colum) Neither cannot be 0, must be atleast 1
                CalledValue = ws.cell(StartWorkRow+row,col+1)
                CalledValue.value = MatchDetail["info"]['participants'][row][CallValues[col]]
            else:
                #ws.cell(Row,Colum) Neither cannot be 0, must be atleast 1
                CalledValue = ws.cell(StartWorkRow+row,col+1)
                Kills = str(MatchDetail["info"]['participants'][row]["kills"])
                Deaths = str(MatchDetail["info"]['participants'][row]["deaths"])
                Assists = str(MatchDetail["info"]['participants'][row]["assists"])
                CalledValue.value = Kills + "/" + Deaths + "/" + Assists
        print("Filled in Red Team Player "+str(row))

def OutputBlue(worksheet,matchDetail,startRow,CallValues):
    ws = worksheet
    StartWorkRow = startRow + 3

    for row in range(5):
        for col in range(len(CallValues)):
            if CallValues[col] != "kills":
                #ws.cell(Row,Colum) Neither cannot be 0, must be atleast 1
                CalledValue = ws.cell(StartWorkRow+row,col+2+len(CallValues))
                CalledValue.value = MatchDetail["info"]['participants'][row+5][CallValues[col]]
            else:
                #ws.cell(Row,Colum) Neither cannot be 0, must be atleast 1
                CalledValue = ws.cell(StartWorkRow+row,col+2+len(CallValues))
                Kills = str(MatchDetail["info"]['participants'][row+5]["kills"])
                Deaths = str(MatchDetail["info"]['participants'][row+5]["deaths"])
                Assists = str(MatchDetail["info"]['participants'][row+5]["assists"])
                CalledValue.value = Kills + "/" + Deaths + "/" + Assists
        print("Filled in Blue Team Player "+str(row))

#Key = open("MainProjectFiles\Key.txt","r")
#Key = Key.read().rstrip("\n")
print("Welcome to the LOL Data Pipeline. This program transfers an account's last 20 'Classic' games to excel for analysis.")
print("--------------------------------------------------------------------------------------------------------------------")
sleep(2)

Key = str(input("What is your Riot API Key?: "))
print("--------------------------------------------------------------------------------------------------------------------")
MyReg = str(input("What is your region {br1, eun1, euw1, jp1, kr, la1, la2, na1, oc1, tr1, or ru}?: "))
AccountName = str(input("What is your Account Name?: "))
print("--------------------------------------------------------------------------------------------------------------------")

try:
    watcher = LolWatcher(Key)
    MyAccount = watcher.summoner.by_name(MyReg, AccountName)
except:
    print("Either your Region and/or Account Name and/or API-Key is worng. Please try again.")
    sleep(5)
    exit()


MyMatches = watcher.match.matchlist_by_puuid(MyReg, MyAccount["puuid"])

wb = Workbook()

CallValues = ["summonerName","championName","teamPosition","kills","totalDamageDealt", "goldEarned", "visionScore","win"]

GameCount = 0
Row = 1
while GameCount <= 17:
    MatchDetail = watcher.match.by_id(MyReg,MyMatches[GameCount])
    if MatchDetail["info"]["gameMode"] == "CLASSIC":
        MakeBoilerPlateCells(wb["Sheet"],Row,MatchDetail)
        OutputRed(wb["Sheet"],MatchDetail,Row,CallValues)
        OutputBlue(wb["Sheet"],MatchDetail,Row,CallValues)
        Row +=8
    print("Current Game Count: "+str(GameCount))
    GameCount +=1

if not os.path.exists(os.getcwd()+"/Output"):
    os.mkdir(str(os.getcwd()+"/Output"))
    try:
        wb.save(str(os.getcwd())+"/Output/" + AccountName + ".xlsx")
        print("Your file has been saved too " + str(os.getcwd())+"/Output/" + AccountName + ".xlsx" )
        sleep(5)
    except:
        print("The File is open, close before running.")
        exit()
else:
    try:
        wb.save(str(os.getcwd())+"/Output/" + AccountName + ".xlsx")
        print("Your file has been saved too " + str(os.getcwd())+"/Output/" + AccountName + ".xlsx" )
        sleep(5)
    except:
        print("The File is open, close before running.")
        exit()
