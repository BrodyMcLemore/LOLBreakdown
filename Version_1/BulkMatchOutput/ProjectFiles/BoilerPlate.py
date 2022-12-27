from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime

def MakeBoilerPlateCells(worksheet,startRow,matchDetail):
    ws = worksheet

    BoilerPlateCellValues = ["Summoner Name", "Champion","Role","K/D/A", "Total Damage delt", "Gold Earned", "Vision Score","Win"]

    ws.merge_cells(start_row=startRow, start_column=1, end_row=startRow, end_column=len(BoilerPlateCellValues))
    RedTeamTitleCell = ws.cell(startRow,1)
    RedTeamTitleCell.value = "Red Team"
    RedTeamTitleCell.alignment = Alignment(horizontal='center')
    RedTeamTitleCell

    ws.merge_cells(start_row=startRow, start_column=len(BoilerPlateCellValues)+2, end_row=startRow, end_column=(len(BoilerPlateCellValues)*2)+1)
    BlueTeamTitleCell = ws.cell(startRow,len(BoilerPlateCellValues)+2)
    BlueTeamTitleCell.value = "Blue Team"
    BlueTeamTitleCell.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=startRow+1, start_column=1, end_row=startRow+1, end_column=(len(BoilerPlateCellValues)*2)+1)
    GameDate = ws.cell(startRow+1,1)
    GameDate.value = datetime.fromtimestamp((matchDetail["info"]['gameStartTimestamp'] / float(1000))).strftime("%m/%d/%Y, %H:%M:%S")
    GameDate.alignment = Alignment(horizontal='center')

    #Red Team
    for i in range(len(BoilerPlateCellValues)):
        cell = ws.cell(startRow+2,i+1)
        cell.value = BoilerPlateCellValues[i]
    #Blue Team
    for i in range(len(BoilerPlateCellValues)):
        cell = ws.cell(startRow+2,i+len(BoilerPlateCellValues)+2)
        cell.value = BoilerPlateCellValues[i]
    #Divider
    for i in range(7):
        cell = ws.cell(startRow+i+1,len(BoilerPlateCellValues)+1)
        cell.fill = PatternFill("solid", start_color = "000000")
