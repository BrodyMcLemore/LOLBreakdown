from riotwatcher import LolWatcher
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
import os

def MakeBoilerPlateCells(worksheet):
    ws = worksheet

    ws.merge_cells("A1:D1")
    MyTeamTitleCell = ws.cell(1,1)
    MyTeamTitleCell.value = "My Team"
    MyTeamTitleCell.alignment = Alignment(horizontal='center')

    ws.merge_cells("F1:I1")
    OpTeamTitleCell = ws.cell(1,6)
    OpTeamTitleCell.value = "Op Team"
    OpTeamTitleCell.alignment = Alignment(horizontal='center')

    BoilerPlateCellValues = ["Summoner Name", "Champion","Role","K/D/A"]
    for i in range(4):
        cell = ws.cell(2,i+1)
        cell.value = BoilerPlateCellValues[i]
    for i in range(4):
        cell = ws.cell(2,i+6)
        cell.value = BoilerPlateCellValues[i]

    for i in range(7):
        cell = ws.cell(i+1,5)
        cell.fill = PatternFill("solid", start_color = "000000")

Key = open("Key.txt","r")
Key = Key.read().rstrip("\n")

MyKey = Key
MyReg = "na1"
watcher = LolWatcher(MyKey)
AccountName = "ChaoticThunder0"
MyAccount = watcher.summoner.by_name(MyReg, AccountName)
MyMatches = watcher.match.matchlist_by_puuid(MyReg, MyAccount["puuid"])
LastMatchID = MyMatches[0]
MatchDetail = watcher.match.by_id(MyReg,LastMatchID)

wb = Workbook()
MakeBoilerPlateCells(wb["Sheet"])

CallValues = ["summonerName","championName","individualPosition","kills"]
for x in range(4):
    if CallValues[x] != "kills":
        for i in range(5):
            ws = wb["Sheet"]
            cell = ws.cell(i+3,x+1)
            cell.value = MatchDetail["info"]['participants'][i][CallValues[x]]
    else:
        for i in range(5):
            ws = wb["Sheet"]
            cell = ws.cell(i+3,x+1)
            Kills = str(MatchDetail["info"]['participants'][i]["kills"])
            Deaths = str(MatchDetail["info"]['participants'][i]["deaths"])
            Assists = str(MatchDetail["info"]['participants'][i]["assists"])
            cell.value = Kills + "/" + Deaths + "/" + Assists

for x in range(4):
    if CallValues[x] != "kills":
        for i in range(5,10):
            ws = wb["Sheet"]
            cell = ws.cell(i-2,x+6)
            cell.value = MatchDetail["info"]['participants'][i][CallValues[x]]
    else:
        for i in range(5,10):
            ws = wb["Sheet"]
            cell = ws.cell(i-2,x+6)
            Kills = str(MatchDetail["info"]['participants'][i]["kills"])
            Deaths = str(MatchDetail["info"]['participants'][i]["deaths"])
            Assists = str(MatchDetail["info"]['participants'][i]["assists"])
            cell.value = Kills + "/" + Deaths + "/" + Assists



wb.save(str(os.getcwd())+"/test.xlsx")
