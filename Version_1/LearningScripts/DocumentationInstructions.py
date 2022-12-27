from riotwatcher import LolWatcher
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os

def MakeBoilerPlate(worksheet):
    ws = worksheet

    ws.merge_cells("A1:D1")
    MyTeamTitleCell = ws.cell(1,1)
    MyTeamTitleCell.value = "My Team"
    MyTeamTitleCell.alignment = Alignment(horizontal='center')

    ws.merge_cells("F1:I1")
    OpTeamTitleCell = ws.cell(1,6)
    OpTeamTitleCell.value = "Op Team"
    OpTeamTitleCell.alignment = Alignment(horizontal='center')


## Boiler-ish Plate Stuff
# Riot Developer key. Generated for 1 daty usage
MyKey = "RGAPI-61431001-085c-4f28-a2aa-f9f4b68b3136"
# Secifying the region to work in
MyReg = "na1"
# Creating the API integration watcher
watcher = LolWatcher(MyKey)
# A global variable that can reference my account, or any account specified.
MyAccount = watcher.summoner.by_name(MyReg, "ChaoticThunder0")

# Pull a list of my past 20 Matches. This is a list of the MatchId referenced off my indiidual PUUID. PUUID is a ID unique to each player regardless of region
MyMatches = watcher.match.matchlist_by_puuid(MyReg, MyAccount["puuid"])
# The last MatchId in my match list.
LastMatchID = MyMatches[0]
# ALL of the game details from my last match, referenced by the MatchId and the regaion. Stored as a BIG dictionary
MatchDetail = watcher.match.by_id(MyReg,LastMatchID)
# ""MatchDetail["info"]['participants'][*Index*][*Attribute*]""
# The 'info' value is the important information. 'participants' references the players. *Index* allows the 0-9 reference of each player. *Attribute* is the speficic keyword to refernce a value.
# This list of possible Attributes are listed in the "AttributeList.txt"

# Making the Output Workbook
wb = Workbook()
ws = wb["Sheet"]
MakeTitles(wb["Sheet"])



wb.save(str(os.getcwd())+"/test.xlsx")
