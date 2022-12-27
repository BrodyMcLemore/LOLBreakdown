from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
import os

def MakeBoilerPlateCells(worksheet,startRow,):
    ws = worksheet

    ws.merge_cells(start_row=startRow, start_column=1, end_row=startRow, end_column=4)
    MyTeamTitleCell = ws.cell(startRow,1)
    MyTeamTitleCell.value = "My Team"
    MyTeamTitleCell.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=startRow, start_column=6, end_row=startRow, end_column=9)
    OpTeamTitleCell = ws.cell(startRow,6)
    OpTeamTitleCell.value = "Op Team"
    OpTeamTitleCell.alignment = Alignment(horizontal='center')

    BoilerPlateCellValues = ["Summoner Name", "Champion","Role","K/D/A"]
    for i in range(4):
        cell = ws.cell(startRow+2,i+1)
        cell.value = BoilerPlateCellValues[i]
    for i in range(4):
        cell = ws.cell(startRow+2,i+6)
        cell.value = BoilerPlateCellValues[i]

    for i in range(7):
        cell = ws.cell(startRow+i+1,5)
        cell.fill = PatternFill("solid", start_color = "000000")

wb = Workbook()
for i in range(1,27,8):
    # Yeilds 4 blocks
    MakeBoilerPlateCells(wb["Sheet"],i)

wb.save(str(os.getcwd())+"/work.xlsx")
