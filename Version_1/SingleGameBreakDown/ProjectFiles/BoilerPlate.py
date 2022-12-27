from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime

RowValues = {
    "Combat": ["KDA","Largest Killing Spree","Largest Multi Kill","Lane"],
    "Damage Dealt": ["Totalt Damage Dealt to Champions","Physical Damage Dealt to Champions","Magic Damage Dealt to Champions","True Damage Dealt to Champions"],
    "Damage Taken and Healed": ["Damage Healed","Damage Taken","Self Mitigated Damage","Taken the Most Damage"],
    "Vision":["Vison Score"	,"Wards Placeded","Wards Destroyed","Control Wards Purchased"],
    "Income & Items": ["Golded Earned","Gold Spent","Items"],
    "Minons":["CS","Neutral Minions Killed","Nuetral Minins Killed in Team Jungle","Nuetral Minins Killed in Enemy Jungle"],
    "Misc":["First Blood","Towers Destroyed","Inhibitors Destroyed","Win"]
}

def MakeHeader(Title,Row,Lenght,WorkSheet):
    ws = WorkSheet
    ws.merge_cells(start_row=Row, start_column=1, end_row=Row, end_column=Lenght)
    MergedCell = ws.cell(Row,1)
    MergedCell.value = Title
    MergedCell.alignment = Alignment(horizontal='center')


def MakeSheets(WorkBook,Title):
    wb = WorkBook
    ws = wb.create_sheet(Title)
    return ws

def MakeBoilerPlateCells(worksheet,startRow,matchDetail):
    ws = worksheet

    BoilerPlateCellValues = ["Summoner Name", "Champion","Role","K/D/A", "Total Damage delt", "Gold Earned", "Vision Score","Win"]

    ws.merge_cells(start_row=startRow, start_column=1, end_row=startRow, end_column=len(BoilerPlateCellValues))
    MyTeamTitleCell = ws.cell(startRow,1)
    MyTeamTitleCell.value = "Red Team"
    MyTeamTitleCell.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=startRow, start_column=len(BoilerPlateCellValues)+2, end_row=startRow, end_column=(len(BoilerPlateCellValues)*2)+1)
    OpTeamTitleCell = ws.cell(startRow,len(BoilerPlateCellValues)+2)
    OpTeamTitleCell.value = "Blue Team"
    OpTeamTitleCell.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=startRow+1, start_column=1, end_row=startRow+1, end_column=(len(BoilerPlateCellValues)*2)+1)
    GameDate = ws.cell(startRow+1,1)
    GameDate.value = datetime.fromtimestamp((matchDetail["info"]['gameStartTimestamp'] / float(1000))).strftime("%m/%d/%Y, %H:%M:%S")
    GameDate.alignment = Alignment(horizontal='center')

    #Red Team
    for i in range(len(BoilerPlateCellValues)):
        cell = ws.cell(startRow+2,i+1)
        cell.value = BoilerPlateCellValues[i]
    #Blue Team
    for i in range(len(BoilerPlateCellValues)):
        cell = ws.cell(startRow+2,i+len(BoilerPlateCellValues)+2)
        cell.value = BoilerPlateCellValues[i]
    #Divider
    for i in range(7):
        cell = ws.cell(startRow+i+1,len(BoilerPlateCellValues)+1)
        cell.fill = PatternFill("solid", start_color = "000000")
